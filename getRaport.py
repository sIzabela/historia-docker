import os
import json
import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine
from openpyxl import load_workbook
from dotenv import load_dotenv
from logFunctions import *

load_dotenv()
with open('config.json') as f:
    config = json.load(f)

def get_today_raport_with_headers(output_path):
    log_message("Pobieranie raportu z bazy danych")

    if os.path.exists(output_path):
        log_message(f"Plik {output_path} już istnieje.")
    else:
        USER = os.getenv('SQL_USER')
        PASSWD = os.getenv('SQL_PASSWD')
        SERVER_NAME = os.getenv('SQL_SERVER')
        DB_NAME = os.getenv('SQL_DATABASE')
        SQL_QUERY = config['sql']['sql_select']
        HEADERS = config['sql']['headers']
        START_COLUMN = config['sql']['start_column']

        # Łączenie z bazą danych
        try:
            connection_string = f"mssql+pyodbc://{USER}:{PASSWD}@{SERVER_NAME}/{DB_NAME}?driver=ODBC+Driver+18+for+SQL+Server&TrustServerCertificate=Yes"
            engine = create_engine(connection_string)
            query = (SQL_QUERY)
            log_message("Połączono z bazą danych")
        except Exception as e:
            log_message(f"Błąd podczas łączenia z bazą danych: {e}")
            return

        # Pobieranie danych z SQL
        try:
            data = pd.read_sql(query, engine)
            log_message("Dane pobrane")
        except Exception as e:
            log_message(f"Błąd podczas pobierania danych: {e}")
            return

        # Eksportowanie danych do pliku Excel
        try:
            data.to_excel(output_path, index=False)
            log_message(f"Dane wyeksportowane do pliku Excel: {output_path}")
        except Exception as e:
            log_message(f"Błąd podczas eksportowania danych do pliku Excel: {e}")
            return
        
        # Dodanie nagłówków do pliku Excel
        # Sprawdzenie, czy plik Excel został poprawnie zapisany
        if os.path.exists(output_path):
            # Dodawanie nagłówków w raporcie
            try:
                workbook = load_workbook(output_path)
                worksheet = workbook.active
                
                for index, header in enumerate(HEADERS, start=START_COLUMN):
                    worksheet.cell(row=1, column=index).value = header

                workbook.save(output_path)
                log_message(f"Nagłówki dodane do kolumn od {START_COLUMN}. w pliku Excel")
            except Exception as e:
                log_message(f"Błąd podczas dodawania nagłówków do kolumn w pliku Excel: {e}")
        else:
            log_message(f"Plik Excel nie został znaleziony: {output_path}")